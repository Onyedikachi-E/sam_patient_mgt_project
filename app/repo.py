from .db_models import PatientARTData, LineListRequest
from typing import Any, Dict, Optional, List
from datetime import datetime, date, timedelta
from sqlalchemy.orm import Session
from fastapi import UploadFile, HTTPException, status
from io import BytesIO
import pandas as pd
from .schemas import PatientARTCreate, LineListRequestResponse
from sqlalchemy import delete
from openpyxl.styles import Border, Side
from openpyxl.styles import Border, Side, Alignment
from openpyxl import load_workbook



# =============================================
# CRUD OPERATIONS
# =============================================
class PatientARTCRUD:
    def __init__(self, db_manager: Session):
        self.db_manager = db_manager

    def parse_date(self, value):
        """
        Convert various Excel/date/string formats to a Python date object.
        Returns None if parsing fails.
        """
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None

        # Already a date/datetime
        if isinstance(value, (date, datetime)):
            return value.date() if isinstance(value, datetime) else value

        # Try as string
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

            # 1. Try your specific known format first: DD/MM/YYYY
            try:
                return datetime.strptime(value, "%d/%m/%Y").date()
            except ValueError:
                pass

            # 2. Other common formats if needed
            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue

            # 3. Try pandas as a last resort (but don't fall back to 1970):
            try:
                dt = pd.to_datetime(value, errors="coerce", dayfirst=True)
                if pd.isna(dt):
                    return None
                return dt.date()
            except Exception:
                return None

        # Excel numeric (serial date)
        # If your Excel stores dates as numbers (e.g. 31607), handle here:
        if isinstance(value, (int, float)) and not pd.isna(value):
            try:
                dt = pd.to_datetime(value, origin="1899-12-30", unit="D")
                return dt.date()
            except Exception:
                return None

        # Anything else: give up gracefully
        return None
    
    def create_patient_from_line_list(self, line_list_data: UploadFile):
        """
        Create patient records from uploaded patient line list
        Required fields: state, lga, facility_name_all, datim_code, 
                        sex, hospital_number, patient_identifier, current_age
        """
        try:
            # Implement the logic
            file_bytes = line_list_data.file.read()
            if not file_bytes:
                raise ValueError("Uploaded file is empty.")
            
            # Here, let's assume the first sheet only; change if needed.
            excel_file = BytesIO(file_bytes)
            dataframe = pd.read_excel(excel_file)

            created_count = 0
            # Helper to read safely from a row
            def get_val(row, col):
                return None if (col not in row or pd.isna(row[col])) else row[col]
            for _, row in dataframe.iterrows():
                # Basic safety: skip if essential identifier missing
                if pd.isna(row["patient_identifier"]):
                    continue
                # Skip if patient record already exist
                patient_exist = self.db_manager.query(
                    PatientARTData
                ).filter(PatientARTData.patient_identifier==get_val(row, "patient_identifier")
                ).first()
                if patient_exist:
                    continue

                patient = PatientARTData(
                    state=str(get_val(row, "state")).strip() if get_val(row, "state") is not None else None,
                    lga=str(get_val(row, "lga")).strip() if get_val(row, "lga") is not None else None,
                    facility_name_all=str(get_val(row, "facility_name_all")).strip() if get_val(row, "facility_name_all") is not None else None,
                    datim_code=str(get_val(row, "datim_code")).strip() if get_val(row, "datim_code") is not None else None,
                    sex=str(get_val(row, "sex")).strip() if get_val(row, "sex") is not None else None,
                    hospital_number=str(get_val(row, "hospital_number")).strip() if get_val(row, "hospital_number") is not None else None,
                    patient_identifier=str(get_val(row, "patient_identifier")).strip() if get_val(row, "patient_identifier") is not None else None,
                    current_age=int(get_val(row, "current_age")) if get_val(row, "current_age") is not None else None,

                    date_of_birth=self.parse_date(get_val(row, "date_of_birth")),
                    care_entry_point=str(get_val(row, "care_entry_point")).strip() if get_val(row, "care_entry_point") is not None else None,
                    art_start_date=self.parse_date(get_val(row, "art_start_date")),
                    age_at_art_initiation=int(get_val(row, "age_at_art_initiation")) if get_val(row, "age_at_art_initiation") is not None else None,
                    clients_current_art_status=str(get_val(row, "clients_current_art_status")).strip() if get_val(row, "clients_current_art_status") is not None else None,
                    educational_status=str(get_val(row, "educational_status")).strip() if get_val(row, "educational_status") is not None else None,
                    residential_address=str(get_val(row, "residential_address")).strip() if get_val(row, "residential_address") is not None else None,
                    last_drug_pick_up_date=self.parse_date(get_val(row, "last_drug_pick_up_date")),
                    last_viral_load_result=str(get_val(row, "last_viral_load_result")).strip() if get_val(row, "last_viral_load_result") is not None else None,
                    cd4_test_cd4_result=str(get_val(row, "cd4_test_cd4_result")).strip() if get_val(row, "cd4_test_cd4_result") is not None else None,
                    adherence_outcome_classification=str(get_val(row, "adherence_outcome_classification")).strip() if get_val(row, "adherence_outcome_classification") is not None else None,
                    marital_status=str(get_val(row, "marital_status")).strip() if get_val(row, "marital_status") is not None else None,
                    employment_status=str(get_val(row, "employment_status")).strip() if get_val(row, "employment_status") is not None else None,
                    no_of_days_of_refills=int(get_val(row, "no_of_days_of_refills")) if get_val(row, "no_of_days_of_refills") is not None else None,
                    who_stage_at_art_start=str(get_val(row, "who_stage_at_art_start")).strip() if get_val(row, "who_stage_at_art_start") is not None else None,
                    last_drug_art_pick_up_date=self.parse_date(get_val(row, "last_drug_art_pick_up_date")),
                    duration_on_art_months=int(get_val(row, "duration_on_art_months")) if get_val(row, "duration_on_art_months") is not None else None,
                    previous_art_regimen=str(get_val(row, "previous_art_regimen")).strip() if get_val(row, "previous_art_regimen") is not None else None,
                    current_art_regimen=str(get_val(row, "current_art_regimen")).strip() if get_val(row, "current_art_regimen") is not None else None,
                    current_art_regimen_line=str(get_val(row, "current_art_regimen_line")).strip() if get_val(row, "current_art_regimen_line") is not None else None,
                    last_viral_load_sample_collection_date=self.parse_date(get_val(row, "last_viral_load_sample_collection_date")),
                    last_viral_load_result_date=self.parse_date(get_val(row, "last_viral_load_result_date")),
                    cd4_test_sample_collection_date=self.parse_date(get_val(row, "cd4_test_sample_collection_date")),
                    cd4_test_result_date=self.parse_date(get_val(row, "cd4_test_result_date"))
                )

                self.db_manager.add(patient)
                created_count += 1
            
            self.db_manager.commit()
            return {
                "patient_data": {
                    "message": "Line list import completed",
                    "total_patient_inserted": created_count
                }
            }
        except Exception as e:
            self.db_manager.rollback()
            print(f"✗ Error creating patient records: {str(e)}")
            raise

    # 1. CREATE - Single patient
    def create_patient(self, patient_payload: PatientARTCreate):
        """Create a single patient record"""
        try:
            # Verify patient with identifer doesnt exist
            patient_exist = self.db_manager.query(PatientARTData).filter(PatientARTData.patient_identifier==patient_payload.patient_identifier).first()
            if patient_exist:
                raise HTTPException(
                    status_code=status.HTTP_409_CONFLICT,
                    detail=f"A patient with same patient identifier already exist"
                )
            # Convert Pydantic model to plain dict (only provided fields)
            patient_data = patient_payload.model_dump(exclude_unset=True)

            patient = PatientARTData(**patient_data)
            self.db_manager.add(patient)
            self.db_manager.commit()
            self.db_manager.refresh(patient)
            return {
                "message": "Patient created successfully"
            }
        except Exception as e:
            self.db_manager.rollback()
            print(f"✗ Error creating patient: {str(e)}")
            raise

    
    # 2. READ - Get patient records
    def get_patient_by_identifier(self, patient_identifier: str) -> Optional[PatientARTData]:
        """Get a single patient by patient_identifier"""
        try:
            query = self.db_manager.query(PatientARTData).filter(
                PatientARTData.patient_identifier == patient_identifier,
                PatientARTData.voided==False
            )
            patient = query.first()
            if not patient:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Specified Patient Identifier not found"
                )
            # Compute and attach current ART status
            patient.clients_current_art_status = self.get_art_outcome(
                last_pickup_date=patient.last_drug_pick_up_date,
                days_of_arv_refill=patient.no_of_days_of_refills,
                ltfu_days=28,
                end_date=date.today(),
            )

            return patient
        except:
            raise

    def get_all_patient_identifiers(self, skip, limit):
        """Get all patient records"""
        try:
            query = (
                self.db_manager
                .query(
                    PatientARTData.state,
                    PatientARTData.datim_code, 
                    PatientARTData.patient_identifier)
                .filter(PatientARTData.voided==False)
                .offset(skip)
                .limit(limit)
            )
            patients = query.all()
            return patients
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Error fetching all patients records -> {str(e)})"
            )
    

    def get_all_patients(self, skip, limit) -> List[PatientARTData]:
        """Get all patient records"""
        try:
            query = self.db_manager.query(PatientARTData).filter(PatientARTData.voided==False).offset(skip).limit(limit)
            
            patients = query.all()
            list_of_patients: List[PatientARTData] = []
            for patient in patients:
                patient.clients_current_art_status = self.get_art_outcome(
                    last_pickup_date=patient.last_drug_pick_up_date,
                    days_of_arv_refill=patient.no_of_days_of_refills,
                    ltfu_days=28,
                    end_date=date.today(),
                )
                list_of_patients.append(patient)
                
            return list_of_patients
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Error fetching all patients records -> {str(e)})"
            )
    

    def get_patients_by_datim_code(self, datim_code: str, skip, limit) -> List[PatientARTData]:
        """Get all patients from a specific facility using datim code"""
        try:
            query = self.db_manager.query(PatientARTData).filter(
                PatientARTData.datim_code == datim_code, PatientARTData.voided==False
            ).offset(skip).limit(limit)
            
            patients = query.all()
            list_of_patients: List[PatientARTData] = []
            for patient in patients:
                patient.clients_current_art_status = self.get_art_outcome(
                    last_pickup_date=patient.last_drug_pick_up_date,
                    days_of_arv_refill=patient.no_of_days_of_refills,
                    ltfu_days=28,
                    end_date=date.today(),
                )
                list_of_patients.append(patient)
                
            return list_of_patients
        finally:
            self.db_manager.close()
    

    def get_patients_by_state(self, state: str, skip, limit) -> List[PatientARTData]:
        """Get all patients from a specific state"""
        try:
            query = self.db_manager.query(PatientARTData).filter(
                PatientARTData.state == state, PatientARTData.voided==False
            ).offset(skip).limit(limit)
            
            patients = query.all()
            list_of_patients: List[PatientARTData] = []
            for patient in patients:
                patient.clients_current_art_status = self.get_art_outcome(
                    last_pickup_date=patient.last_drug_pick_up_date,
                    days_of_arv_refill=patient.no_of_days_of_refills,
                    ltfu_days=28,
                    end_date=date.today(),
                )
                list_of_patients.append(patient)
                
            return list_of_patients
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Error fetching all patients records -> {str(e)})"
            )
    

    # 3. UPDATE - Modify existing patient record
    def update_patient(self, patient_identifier: str, update_data: Dict[str, Any]) -> Optional[PatientARTData]:
        """Update an existing patient record"""
        try:
            patient = self.db_manager.query(PatientARTData).filter(
                PatientARTData.patient_identifier == patient_identifier,
                PatientARTData.voided == False
            ).first()
            
            if not patient:
                print(f"✗ Patient not found: {patient_identifier}")
                return None
            
            # Update fields
            for key, value in update_data.items():
                if hasattr(patient, key):
                    setattr(patient, key, value)
            
            self.db_manager.commit()
            self.db_manager.refresh(patient)
            
            print(f"✓ Patient updated successfully: {patient_identifier}")
            return patient
        
        except Exception as e:
            self.db_manager.rollback()
            print(f"✗ Error updating patient: {str(e)}")
            raise
    

    # 4. SOFT DELETE - Void a patient record
    def soft_delete_patient(self, patient_identifier: str, voided_by: str) -> bool:
        """Soft delete a patient record by setting voided = 1"""
        try:
            patient = self.db_manager.query(PatientARTData).filter(
                PatientARTData.patient_identifier == patient_identifier,
                PatientARTData.voided == False
            ).first()
            
            if not patient:
                print(f"✗ Patient not found or already voided: {patient_identifier}")
                return False
            
            patient.voided = 1
            patient.voided_by = voided_by
            patient.voided_date = datetime.now()
            
            self.db_manager.commit()
            
            print(f"✓ Patient voided successfully: {patient_identifier}")
            return True
        
        except Exception as e:
            self.db_manager.rollback()
            print(f"✗ Error voiding patient: {str(e)}")
            raise
    

    # 5. RESTORE - Unvoid a patient record
    def restore_patient(self, patient_identifier: str) -> bool:
        """Restore a voided patient record"""
        try:
            patient = self.db_manager.query(PatientARTData).filter(
                PatientARTData.patient_identifier == patient_identifier,
            ).first()
            
            if not patient:
                print(f"✗ Patient not found: {patient_identifier}")
                return False
            
            if patient.voided==False:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Patient is not voided"
                )
            
            
            patient.voided = 0
            patient.voided_by = None
            patient.voided_date = None
            
            self.db_manager.commit()
            
            print(f"✓ Patient restored successfully: {patient_identifier}")
            return True
        
        except Exception as e:
            self.db_manager.rollback()
            print(f"✗ Error restoring patient: {str(e)}")
            raise


    def drop_all_patients(self) -> int:
        """
        Delete ALL patient records from patient_art_data table.

        Returns:
            int: number of rows deleted
        """
        try:
            # For SQLAlchemy 1.4+ using the Core delete
            stmt = delete(PatientARTData)
            result = self.db_manager.execute(stmt)
            self.db_manager.commit()

            deleted_count = result.rowcount or 0
            print(f"✓ Deleted {deleted_count} patient records")
            return deleted_count

        except Exception as e:
            self.db_manager.rollback()
            print(f"✗ Error deleting all patient records: {str(e)}")
            raise


    
    def generate_patient_line_list(
            self,  
            datim_code: Optional[str] = None
        ) -> BytesIO:
        """
        Generate patient line list from the database.
        Args:
            datim_code (str, optional): Filter by facility DATIM code
        Returns:
            BytesIO: An in-memory Excel file containing the line list
        """
        try:
            # Base query: only non-voided
            query = (
                self.db_manager
                .query(PatientARTData).filter(PatientARTData.voided == False)
            )

            # Optional filter by datim_code
            if datim_code:
                query = query.filter(PatientARTData.datim_code == datim_code)

            patients_list: List[PatientARTData] = query.all()
            list_of_patients: List[PatientARTData] = []
            for patient in patients_list:
                patient.clients_current_art_status = self.get_art_outcome(
                    last_pickup_date=patient.last_drug_pick_up_date,
                    days_of_arv_refill=patient.no_of_days_of_refills,
                    ltfu_days=28,
                    end_date=date.today(),
                )
                list_of_patients.append(patient)

            # Define the exact columns / headers as in your Excel
            columns = [
                "state",
                "lga",
                "facility_name_all",
                "datim_code",
                "sex",
                "hospital_number",
                "patient_identifier",
                "current_age",
                "date_of_birth",
                "care_entry_point",
                "art_start_date",
                "age_at_art_initiation",
                "clients_current_art_status",
                "educational_status",
                "residential_address",
                "last_drug_pick_up_date",
                "last_viral_load_result",
                "cd4_test_cd4_result",
                "adherence_outcome_classification",
                "marital_status",
                "employment_status",
                "no_of_days_of_refills",
                "who_stage_at_art_start",
                "last_drug_art_pick_up_date",
                "duration_on_art_months",
                "previous_art_regimen",
                "current_art_regimen",
                "current_art_regimen_line",
                "last_viral_load_sample_collection_date",
                "last_viral_load_result_date",
                "cd4_test_sample_collection_date",
                "cd4_test_result_date",
                "date",
                "signature",
                "comment",
                "suggestion",
            ]

            # Build rows from ORM objects
            data = []
            for p in list_of_patients:
                row = {
                    "state": p.state,
                    "lga": p.lga,
                    "facility_name_all": p.facility_name_all,
                    "datim_code": p.datim_code,
                    "sex": p.sex,
                    "hospital_number": p.hospital_number,
                    "patient_identifier": p.patient_identifier,
                    "current_age": p.current_age,
                    "date_of_birth": p.date_of_birth,
                    "care_entry_point": p.care_entry_point,
                    "art_start_date": p.art_start_date,
                    "age_at_art_initiation": p.age_at_art_initiation,
                    "clients_current_art_status": p.clients_current_art_status,
                    "educational_status": p.educational_status,
                    "residential_address": p.residential_address,
                    "last_drug_pick_up_date": p.last_drug_pick_up_date,
                    "last_viral_load_result": p.last_viral_load_result,
                    "cd4_test_cd4_result": p.cd4_test_cd4_result,
                    "adherence_outcome_classification": p.adherence_outcome_classification,
                    "marital_status": p.marital_status,
                    "employment_status": p.employment_status,
                    "no_of_days_of_refills": p.no_of_days_of_refills,
                    "who_stage_at_art_start": p.who_stage_at_art_start,
                    "last_drug_art_pick_up_date": p.last_drug_art_pick_up_date,
                    "duration_on_art_months": p.duration_on_art_months,
                    "previous_art_regimen": p.previous_art_regimen,
                    "current_art_regimen": p.current_art_regimen,
                    "current_art_regimen_line": p.current_art_regimen_line,
                    "last_viral_load_sample_collection_date": p.last_viral_load_sample_collection_date,
                    "last_viral_load_result_date": p.last_viral_load_result_date,
                    "cd4_test_sample_collection_date": p.cd4_test_sample_collection_date,
                    "cd4_test_result_date": p.cd4_test_result_date,
                    "date": getattr(p, "date", None),
                    "signature": p.signature,
                    "comment": p.comment,
                    "suggestion": p.suggestion,
                }
                data.append(row)

            # Create DataFrame with the defined column order
            df = pd.DataFrame(data, columns=columns)

            # Write to an in-memory Excel file
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Patient Line List")
            output.seek(0)

            return output
        except Exception as e:
            print(f"✗ Error generating patient line list: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error generating patient line list -> {str(e)}",
            )
        

    def get_line_list_requests(self, skip, limit) -> List["LineListRequestResponse"]:
        """
        Fetch all line list export requests from the database,
        ordered by most recent request date.
        """
        try:
            line_list_data = (
                self.db_manager
                .query(LineListRequest)
                .order_by(LineListRequest.request_date.desc())
                .offset(skip)
                .limit(limit)
                .all()
            )

            line_list_requests = [
                LineListRequestResponse(
                    request_id=req.request_id,
                    requested_by=req.requested_by_id,
                    request_date=req.request_date.strftime("%Y-%m-%d %H:%M:%S"),
                    request_status=req.request_status,
                )
                for req in line_list_data
            ]

            return line_list_requests
        except Exception as e:
            print(f"✗ Error fetching line list requests: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Error fetching line list requests -> {str(e)}",
            )
        

    def get_art_outcome(
        self,
        last_pickup_date: Optional[date],
        days_of_arv_refill: Optional[int],
        ltfu_days: int = 28,
        end_date: Optional[date] = None,
    ) -> str:
        """
        Python equivalent of MySQL getoutcome function.
        Args:
            last_pickup_date: DATE of last ARV pickup
            days_of_arv_refill: Number of days of ARV refilled at last pickup
            ltfu_days: Grace period in days for LTFU (default 28)
            end_date: Reference date (default: today's date)
        Returns:
            "Active", "InActive", or "" (if last_pickup_date is None)
        """
        # enddate is current date if not provided
        if end_date is None:
            end_date = date.today()

        if last_pickup_date is None:
            return "No last pickup date"

        if days_of_arv_refill is None:
            days_of_arv_refill = 0

        ltfunumber = days_of_arv_refill + ltfu_days
        ltfu_date = last_pickup_date + timedelta(days=ltfunumber)
        daysdiff = (ltfu_date - end_date).days

        # IF(daysdiff >=0,'Active','InActive')
        if daysdiff >= 0:
            return "Active"
        else:
            return "Inactive"
    

    def generate_empty_line_list_template(self) -> BytesIO:
        """
        Generate an empty Excel template for patient line list upload.
        - Adds thin borders around cells from row 1 to row 10.
        - Wraps text in the header row (row 1).
        Returns:
            BytesIO: in-memory Excel file with headers and styling.
        """
        try:
            columns = [
                "state",
                "lga",
                "facility_name_all",
                "datim_code",
                "sex",
                "hospital_number",
                "patient_identifier",
                "current_age",
                "date_of_birth",
                "care_entry_point",
                "art_start_date",
                "age_at_art_initiation",
                "clients_current_art_status",
                "educational_status",
                "residential_address",
                "last_drug_pick_up_date",
                "last_viral_load_result",
                "cd4_test_cd4_result",
                "adherence_outcome_classification",
                "marital_status",
                "employment_status",
                "no_of_days_of_refills",
                "who_stage_at_art_start",
                "last_drug_art_pick_up_date",
                "duration_on_art_months",
                "previous_art_regimen",
                "current_art_regimen",
                "current_art_regimen_line",
                "last_viral_load_sample_collection_date",
                "last_viral_load_result_date",
                "cd4_test_sample_collection_date",
                "cd4_test_result_date",
                "date",
                "signature",
                "comment",
                "suggestion",
            ]

            # Empty DataFrame with headers
            df = pd.DataFrame(columns=columns)

            # First write with pandas
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="Patient Line List Template")

            # Style with openpyxl
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active  # "Patient Line List Template"

            # Thin border
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            max_col = ws.max_column
            start_row = 1
            end_row = 10

            # Apply borders rows 1–10
            for row in range(start_row, end_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.border = border

                    # Wrap header row (row 1)
                    if row == 1:
                        cell.alignment = Alignment(wrap_text=True)

            # Save back to BytesIO
            styled_output = BytesIO()
            wb.save(styled_output)
            styled_output.seek(0)

            return styled_output

        except Exception as e:
            print(f"✗ Error generating empty line list template: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Error generating patient line list template -> {e}",
            )